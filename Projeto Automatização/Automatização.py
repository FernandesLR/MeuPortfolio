#Projeto: automatização 2
#Fazer um programa que pegue dados de uma planilha e jogue dentro de um site.

def iniciar(valorPl, valorFol, valorLi):
    from selenium import webdriver # importa o web driver do selenium
    from webdriver_manager.chrome import ChromeDriverManager # importa o manager que vai gerenciar o webdriver e ver qual a versão mais recente do navegador do Chrome
    from selenium.webdriver.chrome.service import Service # importa o service para para instalar a versão mais recente
    import openpyxl
    from os import path
    from time import sleep



    def EnviarValor(num):
        '''
        Essa função serve para enviar os valores da planilha para o site, recebendo um valor númerico da váriavel "c" para acessar uma
        posição do vetor "xpaths".
        '''

        nav.find_element('xpath', f'{xpaths[num]}').send_keys(celula.value) # pede ao navegador para encontrar um xpath especifico para entregar um valor da planilha
        if coluna == 5 or coluna == 10 or coluna == 13: # passa para a próxima página, verificando em qual posição a coluna da planilha está
            # certificando que os valores necessários já foram enviados para o site
            nav.implicitly_wait(1)
            sleep(1)
            nav.find_element('xpath', '/html/body/section/div/div/button').click() # passa para a próxima página
        if nav.current_url == 'file:///C:/Users/leo1r/Documents/Portfolio/Projeto%20Automatiza%C3%A7%C3%A3o/Site%20de%20teste/Cadastro4.html':
            # aqui verifica se o navegador está na última página para cadastrar um outro produto
            nav.find_element('xpath', '/html/body/section/div/div/button').click()


    planilha = openpyxl.load_workbook('Prodloja.xlsx')  # pega valor digitado dentro da interface e procura no computador


    folha = planilha['Produtos'] # acessa a folha dentro da planilha
    c = 0 # um contador para acessar a posição do xpath
    linha = 2
    qtd = int(2) # quantas linhas devem ser automatizadas


    xpaths = ['/html/body/section/div/div/input[1]', '/html/body/section/div/div/input[2]', '/html/body/section/div/div/input[3]', '/html/body/section/div/div/input[4]', 
            '/html/body/section/div/div/input[5]'] # é uma lista com todos os xpaths necessários para a automatização

    servico = Service(ChromeDriverManager().install()) # instala a versão mais recente do webdriver
    nav = webdriver.Chrome(service=servico)
    nav.get('C:/Users/leo1r/Documents/Portfolio/Projeto Automatização/Site de teste/Cadastro.html') # acessa o site


    # Essas condicionais verificam e ajustam para o número correto já que o sistema sempre registra 1 a menos
    if qtd == 1:
        qtd = 2 # O sistema só registra o primeiro produto e não 2 produtos
    elif qtd >= 2:
        qtd += 1 # faz com que o sistema leia a quantidade exata de linhas colocada pelo usuário


    while linha <= qtd:
        for coluna in range(1, folha.max_column + 1): # vai passando de coluna em coluna da planilha
            celula = folha.cell(row=linha, column=coluna) # acessa a célula com o valor
            EnviarValor(c) # passa o contador para a função
            c += 1
            if c == 5: # se o contador for igual a 5, volta para o número 0, pois 5 é o tamanho total da lista xpaths
                c = 0
        c = 0 # reseta o contador após o loop for ser finalizado
        linha += 1 # passa para próxima linha
    print('\033[92mProdutos cadastrados com sucesso!\033[0m')
    return True
